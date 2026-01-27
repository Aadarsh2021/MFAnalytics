import json
import csv
from datetime import datetime

def parse_csv(filepath):
    """Parse CSV with BOM handling"""
    try:
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            data = []
            for row in reader:
                clean_row = {k.strip().strip('"'): v for k, v in row.items()}
                if any(clean_row.values()):
                    data.append(clean_row)
            return data
    except Exception as e:
        print(f"⚠️ Error: {filepath}: {e}")
        return []

def parse_date(date_str):
    """Parse multiple date formats to YYYY-MM-DD"""
    if not date_str:
        return None
    formats = ['%m-%d-%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y']
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str.strip(), fmt)
            # Keep day=01 for consistency
            return dt.strftime('%Y-%m-01')
        except:
            continue
    return None

def safe_float(val):
    """Convert to float safely"""
    if not val:
        return 0.0
    try:
        return float(str(val).replace(',', '').replace('%', '').strip())
    except:
        return 0.0

print("📂 Loading ALL data files...")

# Load all CSV files
files = {
    'fed': parse_csv('Data_For_Model/FED FUNDS EFFECTIVE RATE.csv'),
    'gdp': parse_csv('Data_For_Model/Real GDP Raw Data.csv'),
    'sp500': parse_csv('Data_For_Model/S&P 500 Historical Data.csv'),
    'bond': parse_csv('Data_For_Model/United States 10-Year Bond Yield Historical Data.csv'),
    'rate': parse_csv('Data_For_Model/USD Real Policy Rate- 1.csv'),
    'vol': parse_csv('Data_For_Model/US CPI Inflation Volatility.csv'),
    'exp': parse_csv('Data_For_Model/US FED Interest Expenses Data.csv'),
    'debt': parse_csv('Data_For_Model/FED DEBT to GDP.csv'),
    'pce': parse_csv('Data_For_Model/PCEPI.csv')
}

for name, data in files.items():
    print(f"✅ {name}: {len(data)} records")

# Try Excel files
try:
    import openpyxl
    print("\n📊 Attempting to load Excel files...")
    wb1 = openpyxl.load_workbook('Data_For_Model/GDT Tables_Q424_EN.xlsx', data_only=True)
    wb2 = openpyxl.load_workbook('Data_For_Model/GDT-Tables_Q325_EN.xlsx', data_only=True)
    print(f"✅ Excel file 1: {len(wb1.sheetnames)} sheets")
    print(f"✅ Excel file 2: {len(wb2.sheetnames)} sheets")
except ImportError:
    print("⚠️ openpyxl not installed - skipping Excel files")
except Exception as e:
    print(f"⚠️ Excel error: {e}")

# Build data by date
print("\n📊 Merging all data sources...")
data_by_date = {}

# FED Funds base
for row in files['fed']:
    date = parse_date(row.get('Date'))
    if date:
        data_by_date[date] = {
            'date': date,
            'repoRate': safe_float(row.get('FEDFUNDS')),
            'cpiInflation': safe_float(row.get('CPIAUCSL')),
            'gdpGrowth': 0,
            'wpiInflation': safe_float(row.get('CPIAUCSL')),
            'gSecYield': 0,
            'forexReserves': 0,
            'inrUsd': 0,
            'bankCredit': 0
        }

# Add each data source
for row in files['gdp']:
    date = parse_date(row.get('observation_date'))
    if date in data_by_date:
        data_by_date[date]['gdpGrowth'] = safe_float(row.get('GDPC1'))

for row in files['sp500']:
    date = parse_date(row.get('Date'))
    if date in data_by_date:
        data_by_date[date]['sp500Close'] = safe_float(row.get('Price') or row.get('Close'))

for row in files['bond']:
    date = parse_date(row.get('Date'))
    if date in data_by_date:
        data_by_date[date]['gSecYield'] = safe_float(row.get('Price') or row.get('Close'))

for row in files['rate']:
    date = parse_date(row.get('Date'))
    if date in data_by_date:
        data_by_date[date]['realPolicyRate'] = safe_float(row.get('USD Real Policy Rate'))

for row in files['debt']:
    date = parse_date(row.get('observation_date') or row.get('Date'))
    if date in data_by_date:
        data_by_date[date]['debtToGDP'] = safe_float(row.get('GFDEGDQ188S') or row.get('Value'))

for row in files['pce']:
    date = parse_date(row.get('observation_date') or row.get('Date'))
    if date in data_by_date:
        data_by_date[date]['pceInflation'] = safe_float(row.get('PCEPI') or row.get('Value'))

# Convert and sort
combined = sorted(data_by_date.values(), key=lambda x: x['date'])

# Save
with open('src/data/usMacroHistorical.json', 'w') as f:
    json.dump(combined, f, indent=2)

print(f"\n✅ Created usMacroHistorical.json: {len(combined)} records")
print(f"📅 {combined[0]['date']} → {combined[-1]['date']}")

# Verify
latest = combined[-1]
print(f"\n🔍 Latest record ({latest['date']}):")
for key in ['repoRate', 'cpiInflation', 'gdpGrowth', 'sp500Close', 'gSecYield', 'realPolicyRate', 'debtToGDP']:
    print(f"  {key}: {latest.get(key, 0)}")
